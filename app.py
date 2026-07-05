import json
import os
import csv
import base64
import binascii
import hmac
import hashlib
import re
import time
from datetime import date, timedelta
from functools import wraps
from io import BytesIO, StringIO

import mysql.connector
from mysql.connector import errorcode
from openpyxl import load_workbook
from dotenv import load_dotenv
from flask import Flask, abort, jsonify, make_response, request, send_from_directory, session
from flask_cors import CORS
from mysql.connector import Error as MySQLError
from werkzeug.exceptions import HTTPException
from werkzeug.middleware.proxy_fix import ProxyFix
from werkzeug.security import check_password_hash, generate_password_hash

load_dotenv()

# Backend Config: Flask API, MySQL connection, CORS, sessions, and static frontend hosting.
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
FRONTEND_DIR = os.path.abspath(os.path.join(BASE_DIR, "..", "frontend"))
CATALOG_SEED_FILE = os.path.join(BASE_DIR, "catalog_seed.json")
cors_origins = [origin.strip() for origin in os.getenv("FRONTEND_ORIGINS", "").split(",") if origin.strip()]

app = Flask(__name__, static_folder=None)
app.secret_key = os.getenv("SECRET_KEY", "change-this-secret-key")
app.wsgi_app = ProxyFix(app.wsgi_app, x_for=1, x_proto=1, x_host=1, x_prefix=1)

is_production = bool(os.getenv("RENDER"))
session_cookie_samesite = os.getenv("SESSION_COOKIE_SAMESITE") or ("None" if is_production else "Lax")
session_cookie_secure = os.getenv("SESSION_COOKIE_SECURE")
if session_cookie_secure is None:
    session_cookie_secure = "true" if is_production else "false"
app.config.update(
    MAX_CONTENT_LENGTH=int(os.getenv("MAX_UPLOAD_MB", "10")) * 1024 * 1024,
    SESSION_COOKIE_HTTPONLY=True,
    SESSION_COOKIE_SAMESITE=session_cookie_samesite,
    SESSION_COOKIE_SECURE=session_cookie_secure.lower() == "true",
    PERMANENT_SESSION_LIFETIME=timedelta(days=int(os.getenv("SESSION_DAYS", "30"))),
)
cors_allowed_headers = ["Content-Type", "Authorization", "x-access-token", "token"]
if cors_origins:
    CORS(app, origins=cors_origins, supports_credentials=True, allow_headers=cors_allowed_headers)
else:
    CORS(app, supports_credentials=True, allow_headers=cors_allowed_headers)
database_ready = False
login_attempts = {}
LOGIN_WINDOW_SECONDS = int(os.getenv("LOGIN_WINDOW_SECONDS", "300"))
LOGIN_MAX_ATTEMPTS = int(os.getenv("LOGIN_MAX_ATTEMPTS", "8"))

if is_production and app.secret_key == "change-this-secret-key":
    raise RuntimeError("SECRET_KEY must be set to a strong random value in production")


def security_headers(response):
    response.headers.setdefault("X-Content-Type-Options", "nosniff")
    response.headers.setdefault("X-Frame-Options", "DENY")
    response.headers.setdefault("Referrer-Policy", "strict-origin-when-cross-origin")
    response.headers.setdefault("Permissions-Policy", "camera=(), microphone=(), geolocation=()")
    if request.is_secure or is_production:
        response.headers.setdefault("Strict-Transport-Security", "max-age=31536000; includeSubDomains")
    return response


app.after_request(security_headers)


def auto_category(name):
    text = str(name or "").upper()
    if re.search(r"\b(SHAMPOO|HAIR|SCALP|TUGAIN|MINTOP|MANDIL|MORR|MINOX|MINSCALP|CONDITIONER|DYE|GREY)\b", text):
        return "Hair Care"
    if re.search(r"\b(SOAP|BAR|FACE|CREAM|GEL|LOTION|SERUM|SUN|SPF|CLEANSER|MOIST|WASH|MASK|LIP|OINT|NAIL|SPRAY|POWDER|ROLL ON)\b", text):
        return "Skin Care"
    return "Medicine"


def load_seed_catalog_names():
    try:
        with open(CATALOG_SEED_FILE, "r", encoding="utf-8") as seed_file:
            names = json.load(seed_file)
    except (OSError, json.JSONDecodeError) as error:
        app.logger.warning("Catalog seed file could not be loaded: %s", error)
        return []

    if not isinstance(names, list):
        app.logger.warning("Catalog seed file must contain a JSON array")
        return []

    unique_names = []
    seen = set()
    for raw_name in names:
        name = str(raw_name or "").strip()
        key = name.casefold()
        if name and key not in seen:
            unique_names.append(name)
            seen.add(key)
    return unique_names


# Database Helpers: all MySQL reads/writes pass through these functions.
def db_config():
    return {
        "host": os.getenv("MYSQL_HOST", "localhost"),
        "port": int(os.getenv("MYSQL_PORT", "3306")),
        "user": os.getenv("MYSQL_USER", "root"),
        "password": os.getenv("MYSQL_PASSWORD", ""),
        "database": os.getenv("MYSQL_DATABASE", "clinic"),
        "autocommit": False,
    }


def get_db():
    config = db_config()
    try:
        return mysql.connector.connect(**config)
    except MySQLError as error:
        if error.errno != errorcode.ER_BAD_DB_ERROR:
            raise
        database = config.pop("database")
        if not re.match(r"^[A-Za-z0-9_]+$", database):
            raise
        conn = mysql.connector.connect(**config)
        cursor = conn.cursor()
        cursor.execute(f"CREATE DATABASE IF NOT EXISTS `{database}` CHARACTER SET utf8mb4 COLLATE utf8mb4_unicode_ci")
        conn.commit()
        cursor.close()
        conn.close()
        return mysql.connector.connect(**db_config())


def query_all(sql, params=None):
    conn = get_db()
    cursor = conn.cursor(dictionary=True)
    cursor.execute(sql, params or ())
    rows = cursor.fetchall()
    cursor.close()
    conn.close()
    return rows


def query_one(sql, params=None):
    rows = query_all(sql, params)
    return rows[0] if rows else None


def execute(sql, params=None):
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute(sql, params or ())
    conn.commit()
    new_id = cursor.lastrowid
    cursor.close()
    conn.close()
    return new_id


def create_tables():
    statements = [
        """
        CREATE TABLE IF NOT EXISTS users (
            id INT AUTO_INCREMENT PRIMARY KEY,
            username VARCHAR(80) NOT NULL UNIQUE,
            password_hash VARCHAR(255) NOT NULL,
            role ENUM('receptionist', 'doctor') NOT NULL,
            full_name VARCHAR(140) NOT NULL,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
        """,
        """
        CREATE TABLE IF NOT EXISTS patients (
            id INT AUTO_INCREMENT PRIMARY KEY,
            patient_id VARCHAR(40) NOT NULL UNIQUE,
            name VARCHAR(160) NOT NULL,
            age INT NOT NULL,
            gender ENUM('Female', 'Male', 'Other') NOT NULL,
            phone VARCHAR(30) NOT NULL,
            date_of_visit DATE NOT NULL,
            location_area VARCHAR(180) NOT NULL,
            main_concern TEXT NOT NULL,
            created_by VARCHAR(80) NOT NULL,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP ON UPDATE CURRENT_TIMESTAMP,
            INDEX idx_patient_search (name, phone, patient_id),
            INDEX idx_patient_gender (gender),
            INDEX idx_visit_date (date_of_visit)
        )
        """,
        """
        CREATE TABLE IF NOT EXISTS appointments (
            id INT AUTO_INCREMENT PRIMARY KEY,
            patient_db_id INT NOT NULL,
            appointment_date DATE NOT NULL,
            appointment_time VARCHAR(20) NOT NULL,
            doctor_name VARCHAR(120) DEFAULT 'Doctor',
            status ENUM('Scheduled', 'Completed', 'Cancelled') DEFAULT 'Scheduled',
            notes TEXT,
            created_by VARCHAR(80) NOT NULL,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY (patient_db_id) REFERENCES patients(id) ON DELETE CASCADE,
            INDEX idx_appointment_date (appointment_date),
            INDEX idx_appointment_patient (patient_db_id)
        )
        """,
        """
        CREATE TABLE IF NOT EXISTS patient_id_sequences (
            sequence_date DATE PRIMARY KEY,
            next_number INT NOT NULL DEFAULT 1,
            updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP ON UPDATE CURRENT_TIMESTAMP
        )
        """,
        """
        CREATE TABLE IF NOT EXISTS prescription_sequences (
            sequence_date DATE PRIMARY KEY,
            next_number INT NOT NULL DEFAULT 1,
            updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP ON UPDATE CURRENT_TIMESTAMP
        )
        """,
        """
        CREATE TABLE IF NOT EXISTS product_catalog (
            id INT AUTO_INCREMENT PRIMARY KEY,
            name VARCHAR(220) NOT NULL,
            category ENUM('Medicine', 'Skin Care', 'Hair Care') NOT NULL DEFAULT 'Medicine',
            default_dose VARCHAR(120),
            default_notes VARCHAR(220),
            created_by VARCHAR(80) NOT NULL,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            UNIQUE KEY unique_catalog_item (name, category),
            INDEX idx_catalog_name (name),
            INDEX idx_catalog_category (category)
        )
        """,
        """
        CREATE TABLE IF NOT EXISTS prescriptions (
            id INT AUTO_INCREMENT PRIMARY KEY,
            prescription_no VARCHAR(40) NOT NULL UNIQUE,
            patient_db_id INT NOT NULL,
            prescription_date DATE NOT NULL,
            follow_up_date DATE,
            medicines JSON,
            skin_products JSON,
            hair_products JSON,
            session_recommended VARCHAR(140),
            session_type VARCHAR(180),
            treatment_notes TEXT,
            receptionist_instructions TEXT,
            created_by VARCHAR(80) NOT NULL,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY (patient_db_id) REFERENCES patients(id) ON DELETE CASCADE,
            INDEX idx_prescription_date (prescription_date),
            INDEX idx_prescription_patient (patient_db_id)
        )
        """,
    ]
    conn = get_db()
    cursor = conn.cursor()
    for statement in statements:
        cursor.execute(statement)
    conn.commit()
    cursor.close()
    conn.close()
    ensure_production_schema()


def ensure_production_schema():
    conn = get_db()
    cursor = conn.cursor()
    database = os.getenv("MYSQL_DATABASE", "clinic")
    for table in ["patients", "prescriptions"]:
        cursor.execute(
            """
            SELECT COUNT(*) FROM INFORMATION_SCHEMA.COLUMNS
            WHERE TABLE_SCHEMA=%s AND TABLE_NAME=%s AND COLUMN_NAME='deleted_at'
            """,
            (database, table),
        )
        exists = cursor.fetchone()[0]
        if exists:
            cursor.execute(f"ALTER TABLE {table} DROP COLUMN deleted_at")
    for statement in [
        "CREATE INDEX idx_patient_created ON patients (created_at)",
        "CREATE INDEX idx_prescription_created ON prescriptions (created_at)",
    ]:
        try:
            cursor.execute(statement)
        except MySQLError as error:
            if error.errno != errorcode.ER_DUP_KEYNAME:
                raise
    conn.commit()
    cursor.close()
    conn.close()


def seed_users():
    users = [
        ("reception", "Reception@123", "receptionist", "Receptionist"),
        ("doctor", "Doctor@123", "doctor", "Doctor"),
    ]
    for username, password, role, full_name in users:
        existing = query_one("SELECT id FROM users WHERE username=%s", (username,))
        if not existing:
            execute(
                "INSERT INTO users (username, password_hash, role, full_name) VALUES (%s, %s, %s, %s)",
                (username, generate_password_hash(password), role, full_name),
            )


def seed_catalog():
    existing = query_one("SELECT COUNT(*) AS count FROM product_catalog")
    if existing and int(existing["count"] or 0) > 0:
        app.logger.info("Catalog seed skipped because product_catalog already has data")
        return

    names = load_seed_catalog_names()
    if not names:
        app.logger.warning("Catalog seed skipped because no backend seed items were found")
        return

    rows = [(name, auto_category(name), "", "", "system") for name in names]
    conn = get_db()
    cursor = conn.cursor()
    cursor.executemany(
        """
        INSERT INTO product_catalog (name, category, default_dose, default_notes, created_by)
        VALUES (%s, %s, %s, %s, %s)
        ON DUPLICATE KEY UPDATE name=VALUES(name)
        """,
        rows,
    )
    conn.commit()
    cursor.close()
    conn.close()


def init_database():
    create_tables()
    seed_users()
    seed_catalog()


# Auth Helpers: JWT/session login and optional role checking for protected API routes.
@app.before_request
def ensure_database_ready():
    global database_ready
    if not database_ready and request.path.startswith("/api/"):
        init_database()
        database_ready = True


@app.errorhandler(MySQLError)
def handle_mysql_error(error):
    app.logger.exception("MySQL error")
    return jsonify({"error": "Database connection failed", "detail": str(error)}), 503


@app.errorhandler(Exception)
def handle_unexpected_error(error):
    if isinstance(error, HTTPException):
        return error
    app.logger.exception("Unexpected error")
    return jsonify({"error": "Server error", "detail": str(error)}), 500


def jwt_b64encode(value):
    return base64.urlsafe_b64encode(value).rstrip(b"=").decode("ascii")


def jwt_b64decode(value):
    padding = "=" * (-len(value) % 4)
    return base64.urlsafe_b64decode((value + padding).encode("ascii"))


def create_jwt(user):
    now = int(time.time())
    payload = {
        "sub": str(user["id"]),
        "username": user["username"],
        "role": user["role"],
        "full_name": user["full_name"],
        "iat": now,
        "exp": now + int(app.config["PERMANENT_SESSION_LIFETIME"].total_seconds()),
    }
    header = {"alg": "HS256", "typ": "JWT"}
    signing_input = ".".join(
        [
            jwt_b64encode(json.dumps(header, separators=(",", ":")).encode("utf-8")),
            jwt_b64encode(json.dumps(payload, separators=(",", ":")).encode("utf-8")),
        ]
    )
    signature = hmac.new(app.secret_key.encode("utf-8"), signing_input.encode("ascii"), hashlib.sha256).digest()
    return f"{signing_input}.{jwt_b64encode(signature)}"


def verify_jwt(token):
    try:
        header_part, payload_part, signature_part = token.split(".")
        signing_input = f"{header_part}.{payload_part}"
        expected_signature = hmac.new(app.secret_key.encode("utf-8"), signing_input.encode("ascii"), hashlib.sha256).digest()
        supplied_signature = jwt_b64decode(signature_part)
        if not hmac.compare_digest(expected_signature, supplied_signature):
            return None
        payload = json.loads(jwt_b64decode(payload_part))
        if int(payload.get("exp", 0)) < int(time.time()):
            return None
        return {
            "id": int(payload["sub"]),
            "username": payload["username"],
            "role": payload["role"],
            "full_name": payload["full_name"],
        }
    except (ValueError, KeyError, TypeError, json.JSONDecodeError, binascii.Error, UnicodeDecodeError):
        return None


def bearer_token_from_request():
    auth_header = request.headers.get("Authorization", "")
    if auth_header.lower().startswith("bearer "):
        return auth_header.split(" ", 1)[1].strip()
    return request.headers.get("x-access-token") or request.headers.get("token")


def current_user():
    token = bearer_token_from_request()
    if token:
        user = verify_jwt(token)
        if user:
            return user
    return session.get("user")


def login_required(role=None):
    def decorator(fn):
        @wraps(fn)
        def wrapper(*args, **kwargs):
            user = current_user()
            if not user:
                return jsonify({"error": "Login required"}), 401
            session["user"] = user
            if role and user["role"] != role:
                return jsonify({"error": "Access denied"}), 403
            return fn(*args, **kwargs)

        return wrapper

    return decorator


def client_key():
    forwarded = request.headers.get("X-Forwarded-For", "")
    return (forwarded.split(",")[0].strip() or request.remote_addr or "unknown")[:80]


def check_login_rate_limit(username):
    now = time.time()
    key = f"{client_key()}:{username.lower()}"
    attempts = [stamp for stamp in login_attempts.get(key, []) if now - stamp < LOGIN_WINDOW_SECONDS]
    if len(attempts) >= LOGIN_MAX_ATTEMPTS:
        return False, key, attempts
    return True, key, attempts


def remember_failed_login(key, attempts):
    attempts.append(time.time())
    login_attempts[key] = attempts


def clear_failed_login(key):
    login_attempts.pop(key, None)


def normalize_json_list(value):
    if isinstance(value, list):
        return value
    return []


def row_dates_to_string(row):
    for key, value in list(row.items()):
        if hasattr(value, "isoformat"):
            row[key] = value.isoformat()
    return row


def parse_json_field(value):
    if isinstance(value, str):
        return json.loads(value or "[]")
    return value or []


def next_sequence_value(cursor, table_name):
    if table_name not in {"patient_id_sequences", "prescription_sequences"}:
        raise ValueError("Invalid sequence table")
    today_value = date.today()
    cursor.execute(
        f"INSERT IGNORE INTO {table_name} (sequence_date, next_number) VALUES (%s, 1)",
        (today_value,),
    )
    cursor.execute(f"SELECT next_number FROM {table_name} WHERE sequence_date=%s FOR UPDATE", (today_value,))
    row = cursor.fetchone()
    number = int(row[0])
    cursor.execute(f"UPDATE {table_name} SET next_number=%s WHERE sequence_date=%s", (number + 1, today_value))
    return today_value, number


def next_patient_id(cursor):
    today_value, number = next_sequence_value(cursor, "patient_id_sequences")
    return f"PT-{today_value.strftime('%Y%m%d')}-{number:06d}"


def next_prescription_no(cursor):
    today_value, number = next_sequence_value(cursor, "prescription_sequences")
    return f"RX-{today_value.strftime('%Y%m%d')}-{number:06d}"


def rows_to_csv(rows, headers):
    output = StringIO()
    writer = csv.DictWriter(output, fieldnames=headers, extrasaction="ignore")
    writer.writeheader()
    writer.writerows(rows)
    return output.getvalue()


def catalog_rows_from_upload(file_storage):
    filename = (file_storage.filename or "").lower()
    rows = []
    if filename.endswith(".xlsx"):
        workbook = load_workbook(BytesIO(file_storage.read()), data_only=True)
        sheet = workbook.active
        headers = [str(cell.value or "").strip().lower().replace(" ", "_") for cell in next(sheet.iter_rows(max_row=1))]
        for values in sheet.iter_rows(min_row=2, values_only=True):
            row = dict(zip(headers, values))
            if row.get("name"):
                rows.append(row)
    else:
        content = file_storage.read().decode("utf-8-sig")
        reader = csv.DictReader(StringIO(content))
        rows = [row for row in reader if row.get("name")]
    return rows


# Frontend Routes: Flask can serve the ready frontend directly when deployed together.
@app.route("/")
def home():
    if not os.path.exists(os.path.join(FRONTEND_DIR, "index.html")):
        return jsonify({"status": "ok", "service": "clinic backend"})
    response = make_response(send_from_directory(FRONTEND_DIR, "index.html"))
    response.headers["Cache-Control"] = "no-store"
    return response


@app.route("/<path:filename>")
def frontend_file(filename):
    if filename.startswith("api/"):
        return jsonify({"error": "API endpoint not found"}), 404
    if not os.path.exists(os.path.join(FRONTEND_DIR, filename)):
        abort(404)
    return send_from_directory(FRONTEND_DIR, filename)


@app.route("/api/health")
def health():
    catalog_count = query_one("SELECT COUNT(*) AS count FROM product_catalog")["count"]
    return jsonify({"status": "ok", "database": os.getenv("MYSQL_DATABASE", "clinic"), "catalog": catalog_count})


# API Routes: authentication, patients, appointments, prescriptions, medicine library, exports, and backups.
@app.route("/api/login", methods=["POST"])
def login():
    data = request.get_json(force=True)
    username = data.get("username", "").strip()
    password = data.get("password", "")
    allowed, attempt_key, attempts = check_login_rate_limit(username)
    if not allowed:
        return jsonify({"error": "Too many login attempts. Try again later."}), 429
    user = query_one("SELECT * FROM users WHERE username=%s", (username,))
    if not user or not check_password_hash(user["password_hash"], password):
        remember_failed_login(attempt_key, attempts)
        return jsonify({"error": "Invalid username or password"}), 401
    clear_failed_login(attempt_key)
    session.permanent = True
    session["user"] = {
        "id": user["id"],
        "username": user["username"],
        "role": user["role"],
        "full_name": user["full_name"],
    }
    return jsonify({"user": session["user"], "token": create_jwt(session["user"])})


@app.route("/api/logout", methods=["POST"])
def logout():
    session.clear()
    return jsonify({"message": "Logged out"})


@app.route("/api/me")
def me():
    return jsonify({"user": current_user()})


@app.route("/api/dashboard")
@login_required()
def dashboard():
    today = date.today().isoformat()
    total_patients = query_one("SELECT COUNT(*) AS count FROM patients")["count"]
    today_patients = query_one("SELECT COUNT(*) AS count FROM patients WHERE date_of_visit=%s", (today,))["count"]
    appointments = query_one("SELECT COUNT(*) AS count FROM appointments WHERE appointment_date=%s", (today,))["count"]
    prescriptions = query_one("SELECT COUNT(*) AS count FROM prescriptions")["count"]
    gender_rows = query_all("SELECT gender, COUNT(*) AS count FROM patients GROUP BY gender")
    patients_by_gender = {"Female": 0, "Male": 0, "Other": 0}
    for row in gender_rows:
        patients_by_gender[row["gender"]] = row["count"]
    total_sessions = query_one(
        """
        SELECT COUNT(*) AS count FROM prescriptions
        WHERE COALESCE(session_recommended, '') <> '' OR COALESCE(session_type, '') <> ''
        """
    )["count"]
    session_rows = query_all(
        """
        SELECT pr.id, pr.prescription_date, pr.follow_up_date, pr.session_recommended, pr.session_type,
               pr.treatment_notes, pr.receptionist_instructions, p.patient_id, p.name AS patient_name
        FROM prescriptions pr
        JOIN patients p ON p.id = pr.patient_db_id
        WHERE COALESCE(pr.session_recommended, '') <> '' OR COALESCE(pr.session_type, '') <> ''
        ORDER BY pr.prescription_date DESC, pr.created_at DESC
        LIMIT 8
        """
    )
    return jsonify(
        {
            "totalPatients": total_patients,
            "todayPatients": today_patients,
            "todayAppointments": appointments,
            "prescriptions": prescriptions,
            "totalSessions": total_sessions,
            "patientsByGender": patients_by_gender,
            "recentSessions": [row_dates_to_string(row) for row in session_rows],
        }
    )


@app.route("/api/patients", methods=["GET", "POST"])
@login_required()
def patients():
    if request.method == "POST":
        if session["user"]["role"] != "receptionist":
            return jsonify({"error": "Only receptionist can register patients"}), 403
        data = request.get_json(force=True)
        required = ["name", "age", "gender", "phone", "date_of_visit", "location_area", "main_concern"]
        missing = [field for field in required if not str(data.get(field, "")).strip()]
        if missing:
            return jsonify({"error": "Missing fields", "fields": missing}), 400
        conn = get_db()
        cursor = conn.cursor()
        try:
            patient_id = next_patient_id(cursor)
            cursor.execute(
                """
                INSERT INTO patients
                (patient_id, name, age, gender, phone, date_of_visit, location_area, main_concern, created_by)
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)
                """,
                (
                    patient_id,
                    data["name"].strip(),
                    int(data["age"]),
                    data["gender"],
                    data["phone"].strip(),
                    data["date_of_visit"],
                    data["location_area"].strip(),
                    data["main_concern"].strip(),
                    session["user"]["username"],
                ),
            )
            new_id = cursor.lastrowid
            conn.commit()
        except Exception:
            conn.rollback()
            raise
        finally:
            cursor.close()
            conn.close()
        return jsonify({"message": "Patient registered", "id": new_id, "patient_id": patient_id}), 201

    search = request.args.get("search", "").strip()
    gender = request.args.get("gender", "").strip()
    sql = "SELECT * FROM patients WHERE 1=1"
    params = []
    if search:
        sql += " AND (name LIKE %s OR phone LIKE %s OR patient_id LIKE %s)"
        like = f"%{search}%"
        params.extend([like, like, like])
    if gender:
        sql += " AND gender=%s"
        params.append(gender)
    sql += " ORDER BY created_at DESC"
    rows = [row_dates_to_string(row) for row in query_all(sql, tuple(params))]
    return jsonify(rows)


@app.route("/api/patients/<int:patient_id>", methods=["PUT", "DELETE"])
@login_required("receptionist")
def patient_detail(patient_id):
    existing = query_one("SELECT id FROM patients WHERE id=%s", (patient_id,))
    if not existing:
        return jsonify({"error": "Patient not found"}), 404
    if request.method == "DELETE":
        conn = get_db()
        cursor = conn.cursor()
        try:
            cursor.execute("DELETE FROM appointments WHERE patient_db_id=%s", (patient_id,))
            cursor.execute("DELETE FROM prescriptions WHERE patient_db_id=%s", (patient_id,))
            cursor.execute("DELETE FROM patients WHERE id=%s", (patient_id,))
            conn.commit()
        except Exception:
            conn.rollback()
            raise
        finally:
            cursor.close()
            conn.close()
        return jsonify({"message": "Patient permanently deleted"})

    data = request.get_json(force=True)
    required = ["name", "age", "gender", "phone", "date_of_visit", "location_area", "main_concern"]
    missing = [field for field in required if not str(data.get(field, "")).strip()]
    if missing:
        return jsonify({"error": "Missing fields", "fields": missing}), 400
    execute(
        """
        UPDATE patients
        SET name=%s, age=%s, gender=%s, phone=%s, date_of_visit=%s, location_area=%s, main_concern=%s
        WHERE id=%s
        """,
        (
            data["name"].strip(),
            int(data["age"]),
            data["gender"],
            data["phone"].strip(),
            data["date_of_visit"],
            data["location_area"].strip(),
            data["main_concern"].strip(),
            patient_id,
        ),
    )
    return jsonify({"message": "Patient updated", "id": patient_id})


@app.route("/api/appointments", methods=["GET", "POST"])
@login_required("receptionist")
def appointments():
    if request.method == "POST":
        data = request.get_json(force=True)
        new_id = execute(
            """
            INSERT INTO appointments
            (patient_db_id, appointment_date, appointment_time, doctor_name, status, notes, created_by)
            VALUES (%s, %s, %s, %s, %s, %s, %s)
            """,
            (
                int(data["patient_db_id"]),
                data["appointment_date"],
                data["appointment_time"],
                data.get("doctor_name", "Doctor"),
                data.get("status", "Scheduled"),
                data.get("notes", ""),
                session["user"]["username"],
            ),
        )
        return jsonify({"message": "Appointment saved", "id": new_id}), 201

    rows = query_all(
        """
        SELECT a.*, p.patient_id, p.name, p.phone
        FROM appointments a
        JOIN patients p ON p.id = a.patient_db_id
        ORDER BY a.appointment_date DESC, a.appointment_time DESC
        """
    )
    return jsonify([row_dates_to_string(row) for row in rows])


@app.route("/api/prescriptions", methods=["GET", "POST"])
@login_required()
def prescriptions():
    if request.method == "POST":
        if session["user"]["role"] != "doctor":
            return jsonify({"error": "Only doctor can save prescriptions"}), 403
        data = request.get_json(force=True)
        patient_db_id = int(data["patient_db_id"])
        conn = get_db()
        cursor = conn.cursor()
        try:
            prescription_no = next_prescription_no(cursor)
            cursor.execute(
                """
                INSERT INTO prescriptions
                (prescription_no, patient_db_id, prescription_date, follow_up_date, medicines, skin_products,
                 hair_products, session_recommended, session_type, treatment_notes, receptionist_instructions, created_by)
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                """,
                (
                    prescription_no,
                    patient_db_id,
                    data["prescription_date"],
                    data.get("follow_up_date") or None,
                    json.dumps(normalize_json_list(data.get("medicines"))),
                    json.dumps(normalize_json_list(data.get("skin_products"))),
                    json.dumps(normalize_json_list(data.get("hair_products"))),
                    data.get("session_recommended", ""),
                    data.get("session_type", ""),
                    data.get("treatment_notes", ""),
                    data.get("receptionist_instructions", ""),
                    session["user"]["username"],
                ),
            )
            new_id = cursor.lastrowid
            conn.commit()
        except Exception:
            conn.rollback()
            raise
        finally:
            cursor.close()
            conn.close()
        return jsonify({"message": "Prescription saved", "id": new_id, "prescription_no": prescription_no}), 201

    rows = query_all(
        """
        SELECT pr.*, p.patient_id, p.name AS patient_name, p.age, p.gender, p.phone, p.location_area, p.main_concern
        FROM prescriptions pr
        JOIN patients p ON p.id = pr.patient_db_id
        ORDER BY pr.created_at DESC
        """
    )
    for row in rows:
        row_dates_to_string(row)
        for field in ["medicines", "skin_products", "hair_products"]:
            row[field] = parse_json_field(row[field])
    return jsonify(rows)


@app.route("/api/sessions")
@login_required()
def sessions():
    search = request.args.get("search", "").strip()
    params = []
    sql = """
        SELECT pr.*, p.patient_id, p.name AS patient_name, p.age, p.gender, p.phone, p.location_area, p.main_concern
        FROM prescriptions pr
        JOIN patients p ON p.id = pr.patient_db_id
        WHERE 1=1
    """
    if search:
        sql += " AND (p.name LIKE %s OR p.patient_id LIKE %s)"
        like = f"%{search}%"
        params.extend([like, like])
    sql += " ORDER BY pr.prescription_date DESC, pr.created_at DESC"
    rows = query_all(sql, tuple(params))
    for row in rows:
        row_dates_to_string(row)
        for field in ["medicines", "skin_products", "hair_products"]:
            row[field] = parse_json_field(row[field])
    return jsonify(rows)


@app.route("/api/sessions/<int:session_id>", methods=["PUT", "DELETE"])
@login_required("receptionist")
def session_detail(session_id):
    existing = query_one("SELECT id FROM prescriptions WHERE id=%s", (session_id,))
    if not existing:
        return jsonify({"error": "Session not found"}), 404
    if request.method == "DELETE":
        execute("DELETE FROM prescriptions WHERE id=%s", (session_id,))
        return jsonify({"message": "Session permanently deleted", "id": session_id})

    data = request.get_json(force=True)
    execute(
        """
        UPDATE prescriptions
        SET follow_up_date=%s, session_recommended=%s, session_type=%s,
            treatment_notes=%s, receptionist_instructions=%s
        WHERE id=%s
        """,
        (
            data.get("follow_up_date") or None,
            data.get("session_recommended", ""),
            data.get("session_type", ""),
            data.get("treatment_notes", ""),
            data.get("receptionist_instructions", ""),
            session_id,
        ),
    )
    return jsonify({"message": "Session updated", "id": session_id})


@app.route("/api/prescriptions/<int:prescription_id>", methods=["PUT", "DELETE"])
@login_required("doctor")
def prescription_detail(prescription_id):
    existing = query_one("SELECT id, prescription_no FROM prescriptions WHERE id=%s", (prescription_id,))
    if not existing:
        return jsonify({"error": "Prescription not found"}), 404
    if request.method == "DELETE":
        execute("DELETE FROM prescriptions WHERE id=%s", (prescription_id,))
        return jsonify({"message": "Prescription permanently deleted"})

    data = request.get_json(force=True)
    execute(
        """
        UPDATE prescriptions
        SET patient_db_id=%s, prescription_date=%s, follow_up_date=%s, medicines=%s, skin_products=%s,
            hair_products=%s, session_recommended=%s, session_type=%s, treatment_notes=%s,
            receptionist_instructions=%s
        WHERE id=%s
        """,
        (
            int(data["patient_db_id"]),
            data["prescription_date"],
            data.get("follow_up_date") or None,
            json.dumps(normalize_json_list(data.get("medicines"))),
            json.dumps(normalize_json_list(data.get("skin_products"))),
            json.dumps(normalize_json_list(data.get("hair_products"))),
            data.get("session_recommended", ""),
            data.get("session_type", ""),
            data.get("treatment_notes", ""),
            data.get("receptionist_instructions", ""),
            prescription_id,
        ),
    )
    return jsonify({"message": "Prescription updated", "id": prescription_id, "prescription_no": existing["prescription_no"]})


@app.route("/api/reports")
@login_required()
def reports():
    gender_rows = query_all("SELECT gender, COUNT(*) AS count FROM patients GROUP BY gender")
    concern_rows = query_all(
        "SELECT main_concern, COUNT(*) AS count FROM patients GROUP BY main_concern ORDER BY count DESC LIMIT 8"
    )
    return jsonify({"byGender": gender_rows, "topConcerns": concern_rows})


@app.route("/api/catalog", methods=["GET", "POST"])
@login_required()
def catalog():
    if request.method == "POST":
        data = request.get_json(force=True)
        name = data.get("name", "").strip()
        if not name:
            return jsonify({"error": "Name is required"}), 400
        execute(
            """
            INSERT INTO product_catalog (name, category, default_dose, default_notes, created_by)
            VALUES (%s, %s, %s, %s, %s)
            ON DUPLICATE KEY UPDATE default_dose=VALUES(default_dose), default_notes=VALUES(default_notes)
            """,
            (
                name,
                data.get("category", "Medicine"),
                data.get("default_dose", ""),
                data.get("default_notes", ""),
                session["user"]["username"],
            ),
        )
        return jsonify({"message": "Catalog item saved"}), 201

    search = request.args.get("search", "").strip()
    params = []
    sql = "SELECT * FROM product_catalog WHERE 1=1"
    if search:
        sql += " AND name LIKE %s"
        params.append(f"%{search}%")
    sql += " ORDER BY name ASC"
    rows = [row_dates_to_string(row) for row in query_all(sql, tuple(params))]
    return jsonify(rows)


@app.route("/api/catalog/upload", methods=["POST"])
@login_required()
def upload_catalog():
    uploaded = request.files.get("file")
    if not uploaded:
        return jsonify({"error": "Upload CSV or XLSX file"}), 400
    rows = catalog_rows_from_upload(uploaded)
    saved = 0
    for row in rows:
        name = str(row.get("name") or row.get("medicine") or row.get("product") or "").strip()
        if not name:
            continue
        category = str(row.get("category") or "Medicine").strip()
        if category.lower().startswith("skin"):
            category = "Skin Care"
        elif category.lower().startswith("hair"):
            category = "Hair Care"
        else:
            category = "Medicine"
        execute(
            """
            INSERT INTO product_catalog (name, category, default_dose, default_notes, created_by)
            VALUES (%s, %s, %s, %s, %s)
            ON DUPLICATE KEY UPDATE default_dose=VALUES(default_dose), default_notes=VALUES(default_notes)
            """,
            (
                name,
                category,
                str(row.get("dose") or row.get("default_dose") or ""),
                str(row.get("notes") or row.get("default_notes") or ""),
                session["user"]["username"],
            ),
        )
        saved += 1
    return jsonify({"message": "Catalog uploaded", "saved": saved})


@app.route("/api/catalog/<int:item_id>", methods=["PUT", "DELETE"])
@login_required()
def catalog_detail(item_id):
    existing = query_one("SELECT id FROM product_catalog WHERE id=%s", (item_id,))
    if not existing:
        return jsonify({"error": "Catalog item not found"}), 404
    if request.method == "DELETE":
        execute("DELETE FROM product_catalog WHERE id=%s", (item_id,))
        return jsonify({"message": "Catalog item deleted"})

    data = request.get_json(force=True)
    name = data.get("name", "").strip()
    if not name:
        return jsonify({"error": "Name is required"}), 400
    category = data.get("category", "Medicine")
    if category not in ["Medicine", "Skin Care", "Hair Care"]:
        category = "Medicine"
    execute(
        """
        UPDATE product_catalog
        SET name=%s, category=%s, default_dose=%s, default_notes=%s
        WHERE id=%s
        """,
        (
            name,
            category,
            data.get("default_dose", ""),
            data.get("default_notes", ""),
            item_id,
        ),
    )
    return jsonify({"message": "Catalog item updated", "id": item_id})


@app.route("/api/export/<kind>")
@login_required()
def export_csv(kind):
    if kind == "patients":
        rows = [row_dates_to_string(row) for row in query_all("SELECT * FROM patients ORDER BY created_at DESC")]
        csv_text = rows_to_csv(rows, ["patient_id", "name", "age", "gender", "phone", "date_of_visit", "location_area", "main_concern", "created_at"])
    elif kind == "prescriptions":
        rows = [row_dates_to_string(row) for row in query_all("SELECT * FROM prescriptions ORDER BY created_at DESC")]
        csv_text = rows_to_csv(rows, ["prescription_no", "patient_db_id", "prescription_date", "follow_up_date", "medicines", "skin_products", "hair_products", "session_recommended", "session_type", "treatment_notes", "receptionist_instructions", "created_at"])
    elif kind == "sessions":
        rows = [row_dates_to_string(row) for row in query_all("SELECT prescription_no, prescription_date, follow_up_date, session_recommended, session_type, treatment_notes FROM prescriptions ORDER BY prescription_date DESC")]
        csv_text = rows_to_csv(rows, ["prescription_no", "prescription_date", "follow_up_date", "session_recommended", "session_type", "treatment_notes"])
    else:
        return jsonify({"error": "Unknown export"}), 404
    return app.response_class(csv_text, mimetype="text/csv", headers={"Content-Disposition": f"attachment; filename=clinic-{kind}.csv"})


@app.route("/api/backup")
@login_required()
def backup():
    data = {
        "patients": [row_dates_to_string(row) for row in query_all("SELECT * FROM patients")],
        "appointments": [row_dates_to_string(row) for row in query_all("SELECT * FROM appointments")],
        "prescriptions": [row_dates_to_string(row) for row in query_all("SELECT * FROM prescriptions")],
        "catalog": [row_dates_to_string(row) for row in query_all("SELECT * FROM product_catalog")],
    }
    output = StringIO()
    writer = csv.DictWriter(output, fieldnames=["table", "payload"])
    writer.writeheader()
    for table, rows in data.items():
        for row in rows:
            writer.writerow({"table": table, "payload": json.dumps(row, default=str)})
    return app.response_class(output.getvalue(), mimetype="text/csv", headers={"Content-Disposition": "attachment; filename=clinic-backup.csv"})


if __name__ == "__main__":
    init_database()
    app.run(host="0.0.0.0", port=int(os.getenv("PORT", "5000")), debug=os.getenv("FLASK_DEBUG") == "1")
